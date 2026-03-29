import cv2
import csv
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime
import numpy as np
from ultralytics import YOLO
from collections import defaultdict

def ccw(A, B, C):
    return (C[1] - A[1]) * (B[0] - A[0]) > (B[1] - A[1]) * (C[0] - A[0])

def intersect(A, B, C, D):
    return ccw(A, C, D) != ccw(B, C, D) and ccw(A, B, C) != ccw(A, B, D)

# ตัวแปร Global
line_pts = []
current_mouse_pos = None
stop_counting = False
manual_reload = False

def draw_line_cb(event, x, y, flags, param):
    global line_pts, current_mouse_pos
    if event == cv2.EVENT_MOUSEMOVE:
        current_mouse_pos = (x, y)
    elif event == cv2.EVENT_LBUTTONDOWN:
        if len(line_pts) == 2:
            line_pts = []
        line_pts.append((x, y))

def click_stop_button(event, x, y, flags, param):
    global stop_counting, manual_reload
    if event == cv2.EVENT_LBUTTONDOWN:
        btn_x1, btn_y1, btn_x2, btn_y2, rld_x1, rld_y1, rld_x2, rld_y2 = param
        if btn_x1 <= x <= btn_x2 and btn_y1 <= y <= btn_y2:
            stop_counting = True
        if rld_x1 <= x <= rld_x2 and rld_y1 <= y <= rld_y2:
            manual_reload = True

# =========================================
# ฟังก์ชันช่วยวาด UI บน OpenCV
# =========================================

def draw_rounded_rect(img, pt1, pt2, color, radius=12, thickness=-1):
    x1, y1 = pt1
    x2, y2 = pt2
    cv2.rectangle(img, (x1 + radius, y1), (x2 - radius, y2), color, thickness)
    cv2.rectangle(img, (x1, y1 + radius), (x2, y2 - radius), color, thickness)
    cv2.circle(img, (x1 + radius, y1 + radius), radius, color, thickness)
    cv2.circle(img, (x2 - radius, y1 + radius), radius, color, thickness)
    cv2.circle(img, (x1 + radius, y2 - radius), radius, color, thickness)
    cv2.circle(img, (x2 - radius, y2 - radius), radius, color, thickness)

def draw_text_centered(img, text, center_x, y, font, scale, color, thickness=1):
    (tw, _), _ = cv2.getTextSize(text, font, scale, thickness)
    cv2.putText(img, text, (center_x - tw // 2, y), font, scale, color, thickness, cv2.LINE_AA)

# สีเส้นนับแต่ละเส้น (BGR)
LINE_COLORS = [
    (0, 220, 255),   # Cyan   — Line 1
    (50, 235, 100),  # Green  — Line 2
    (0, 165, 255),   # Orange — Line 3
]
LINE_LABEL_COLORS_HEX = ["#00dcff", "#32eb64", "#00a5ff"]  # for Tkinter

# ==========================================
# ฟังก์ชันหลักนับยานพาหนะ
# ==========================================

def run_counter(source, export_path, fast_mode, num_lines, line_labels, model_size="yolov8n.pt", conf_thresh=0.20, iou_thresh=0.50):
    global line_pts, stop_counting, manual_reload

    print(f"กำลังโหลดโมเดล: {model_size} และ Source: {source}")
    model = YOLO(model_size)

    vehicle_classes = [2, 3, 5, 7]
    class_names = model.names

    FONT = cv2.FONT_HERSHEY_SIMPLEX

    def connect_stream():
        nonlocal cap
        if 'cap' in locals() and cap is not None:
            cap.release()
        cap_source = int(source) if source.isdigit() else source
        cap = cv2.VideoCapture(cap_source)
        cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
        return cap

    cap = None
    cap = connect_stream()

    if not cap.isOpened():
        messagebox.showerror("Error", f"ไม่สามารถเปิด Video หรือ Streaming นี้ได้:\n{source}")
        return

    ret, frame = cap.read()
    if not ret:
        messagebox.showerror("Error", "อ่านเฟรมแรกของวิดีโอไม่ได้")
        cap.release()
        return

    h, w = frame.shape[:2]
    max_height = 480
    if h > max_height:
        scale = max_height / h
        new_w, new_h = int(w * scale), int(h * scale)
    else:
        new_w, new_h = w, h

    frame = cv2.resize(frame, (new_w, new_h))

    # ==========================================
    # ส่วนที่ 1: วาดเส้นนับทีละเส้น
    # ==========================================
    cv2.namedWindow("Setup Line")
    cv2.setMouseCallback("Setup Line", draw_line_cb)

    lines = []  # เก็บ (start, end) ของแต่ละเส้น

    for ln_idx in range(num_lines):
        line_pts = []
        current_mouse_pos = None
        label = line_labels[ln_idx]
        color = LINE_COLORS[ln_idx]
        header_text = f"SETUP LINE {ln_idx + 1} / {num_lines}  —  {label}"

        while True:
            temp = frame.copy()

            # วาดเส้นที่ confirm แล้วก่อนหน้า
            for prev_i, (ps, pe) in enumerate(lines):
                pc = LINE_COLORS[prev_i]
                cv2.line(temp, ps, pe, pc, 2, cv2.LINE_AA)
                cv2.circle(temp, ps, 5, (255, 255, 255), -1)
                cv2.circle(temp, pe, 5, (255, 255, 255), -1)
                cv2.putText(temp, line_labels[prev_i], (ps[0] + 5, ps[1] - 10),
                            FONT, 0.45, pc, 1, cv2.LINE_AA)

            # วาดเส้นปัจจุบันที่กำลังวาด
            if len(line_pts) >= 1:
                cv2.circle(temp, line_pts[0], 7, color, -1)
                cv2.circle(temp, line_pts[0], 9, (255, 255, 255), 1)
            if len(line_pts) == 1 and current_mouse_pos is not None:
                cv2.line(temp, line_pts[0], current_mouse_pos, color, 1, cv2.LINE_AA)
            if len(line_pts) == 2:
                cv2.line(temp, line_pts[0], line_pts[1], color, 2, cv2.LINE_AA)
                cv2.circle(temp, line_pts[1], 7, color, -1)
                cv2.circle(temp, line_pts[1], 9, (255, 255, 255), 1)
                cv2.putText(temp, label, (line_pts[0][0] + 5, line_pts[0][1] - 12),
                            FONT, 0.5, color, 1, cv2.LINE_AA)

            # Hint bar ด้านล่าง
            ov = temp.copy()
            cv2.rectangle(ov, (0, new_h - 42), (new_w, new_h), (15, 15, 15), -1)
            cv2.addWeighted(ov, 0.75, temp, 0.25, 0, temp)
            draw_text_centered(temp,
                               "Click 2 points to draw  |  [C] Confirm  |  [Q] Quit",
                               new_w // 2, new_h - 14, FONT, 0.45, (200, 200, 200))

            # Header bar
            ov2 = temp.copy()
            cv2.rectangle(ov2, (0, 0), (new_w, 38), (20, 20, 20), -1)
            cv2.addWeighted(ov2, 0.8, temp, 0.2, 0, temp)
            draw_text_centered(temp, header_text, new_w // 2, 25, FONT, 0.58, color, 2)

            cv2.imshow("Setup Line", temp)
            key = cv2.waitKey(1) & 0xFF

            if key == ord('c'):
                if len(line_pts) == 2:
                    lines.append((line_pts[0], line_pts[1]))
                    break
                else:
                    print(f"กรุณาคลิกให้ครบ 2 จุดก่อนกด C (Line {ln_idx + 1})")
            elif key == ord('q'):
                cap.release()
                cv2.destroyAllWindows()
                return

    cv2.destroyWindow("Setup Line")

    # ==========================================
    # ส่วนที่ 2: เริ่มนับยานพาหนะ
    # ==========================================
    track_history = defaultdict(list)

    # แต่ละเส้นมี counted_ids, counts, export_data เป็นของตัวเอง
    counted_ids_per_line = [set() for _ in range(num_lines)]
    counts_per_line = [
        {"inbound": defaultdict(int), "outbound": defaultdict(int)}
        for _ in range(num_lines)
    ]
    export_data_per_line = [[] for _ in range(num_lines)]

    cv2.namedWindow("Vehicle Counting")
    cv2.namedWindow("Dashboard")

    # --- Dashboard ขนาดปรับตามจำนวนเส้น ---
    rows_per_line = 5   # header + 4 vehicle classes max
    px_per_row    = 24
    section_h     = 14 + rows_per_line * px_per_row + 18
    dash_w        = 400
    dash_content_h= 65 + 44 + num_lines * section_h
    btn_area_h    = 44 + 56 + 24
    dash_h        = dash_content_h + btn_area_h + 30

    btn_w, btn_h = 170, 44
    btn_x1 = (dash_w - btn_w) // 2
    btn_y1 = dash_h - btn_h - 18
    btn_x2 = btn_x1 + btn_w
    btn_y2 = btn_y1 + btn_h

    rld_w, rld_h = 170, 44
    rld_x1 = btn_x1
    rld_y1 = btn_y1 - rld_h - 12
    rld_x2 = rld_x1 + rld_w
    rld_y2 = rld_y1 + rld_h

    cv2.setMouseCallback("Dashboard", click_stop_button,
                         param=(btn_x1, btn_y1, btn_x2, btn_y2, rld_x1, rld_y1, rld_x2, rld_y2))

    last_frame_time = datetime.now()

    # --- สี Palette ---
    BG_COLOR       = (18, 18, 28)
    HEADER_COLOR   = (28, 28, 45)
    ACCENT         = (0, 200, 255)
    TEXT_COLOR     = (220, 220, 230)
    MUTED_COLOR    = (110, 110, 130)
    BTN_STOP       = (40, 50, 200)
    BTN_RELOAD     = (20, 140, 200)
    DIVIDER_COLOR  = (40, 40, 60)
    INBOUND_COLOR  = (80, 220, 130)
    OUTBOUND_COLOR = (80, 130, 255)

    stop_counting = False
    manual_reload = False
    
    frame_count = 0
    last_dashboard = None

    while True:
        if stop_counting:
            break

        if not cap.isOpened():
            cap = connect_stream()
            if not cap.isOpened():
                cv2.waitKey(1000)
                continue

        ret, frame = cap.read()
        time_since_last = (datetime.now() - last_frame_time).total_seconds()
        is_stalled = (not ret and not str(source).isdigit()) or \
                     (time_since_last > 5.0 and not str(source).isdigit())

        if is_stalled or manual_reload:
            manual_reload = False
            cap = connect_stream()
            last_frame_time = datetime.now()
            continue

        if not ret:
            if str(source).isdigit() or source.lower().endswith(('.mp4', '.avi', '.mov', '.mkv')):
                print("สิ้นสุดไฟล์วิดีโอ")
                break
            continue

        last_frame_time = datetime.now()
        frame = cv2.resize(frame, (new_w, new_h))
        frame_count += 1

        tracking_imgsz = 320 if fast_mode else 640
        results = model.track(frame, persist=True, classes=vehicle_classes,
                               imgsz=tracking_imgsz, conf=conf_thresh, iou=iou_thresh, verbose=False)

        # วาดเส้นทุกเส้นบน frame
        for li, (ls, le) in enumerate(lines):
            lc = LINE_COLORS[li]
            cv2.line(frame, ls, le, lc, 2, cv2.LINE_AA)
            cv2.circle(frame, ls, 5, (255, 255, 255), -1)
            cv2.circle(frame, le, 5, (255, 255, 255), -1)
            cv2.putText(frame, line_labels[li], (ls[0] + 6, ls[1] - 8),
                        FONT, 0.4, lc, 1, cv2.LINE_AA)

        if results[0].boxes.id is not None:
            boxes     = results[0].boxes.xyxy.cpu()
            track_ids = results[0].boxes.id.int().cpu().tolist()
            class_ids = results[0].boxes.cls.int().cpu().tolist()

            for box, track_id, class_id in zip(boxes, track_ids, class_ids):
                x1, y1, x2, y2 = map(int, box)
                cx, cy = (x1 + x2) // 2, (y1 + y2) // 2

                hist = track_history[track_id]
                hist.append((cx, cy))
                if len(hist) > 30:
                    hist.pop(0)

                cls_name = class_names[class_id]
                label = f"{cls_name}  #{track_id}"

                cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 220, 255), 1)
                (lw, lh), _ = cv2.getTextSize(label, FONT, 0.4, 1)
                cv2.rectangle(frame, (x1, y1 - lh - 8), (x1 + lw + 6, y1), (0, 220, 255), -1)
                cv2.putText(frame, label, (x1 + 3, y1 - 4),
                            FONT, 0.4, (10, 10, 10), 1, cv2.LINE_AA)

                # เร่งการวาด trail โดยลดความถี่ในการวาดถ้าเฟรมตกค้าง
                for i in range(max(1, len(hist)-10), len(hist)):
                    alpha = int(255 * i / len(hist))
                    cv2.line(frame, hist[i-1], hist[i], (alpha // 2, alpha, 200), 1, cv2.LINE_AA)
                cv2.circle(frame, (cx, cy), 4, (255, 255, 255), -1)

                if len(hist) >= 2:
                    prev_pt = hist[-2]
                    curr_pt = hist[-1]
                    # ตรวจแต่ละเส้น
                    for li, (ls, le) in enumerate(lines):
                        if track_id not in counted_ids_per_line[li]:
                            if intersect(ls, le, prev_pt, curr_pt):
                                counted_ids_per_line[li].add(track_id)
                                is_inbound = ccw(ls, le, curr_pt)
                                direction = "Inbound" if is_inbound else "Outbound"
                                if is_inbound:
                                    counts_per_line[li]["inbound"][cls_name] += 1
                                else:
                                    counts_per_line[li]["outbound"][cls_name] += 1
                                ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                export_data_per_line[li].append(
                                    [ts, track_id, cls_name, direction])

        # ปรับปรุง: คาดการณ์ Grand Total (ใช้สำหรับมุมซ้ายบนของจอด้วย)
        grand_in  = sum(sum(c["inbound"].values())  for c in counts_per_line)
        grand_out = sum(sum(c["outbound"].values()) for c in counts_per_line)
        grand_total = grand_in + grand_out

        # ====================
        # วาด Dashboard (วาดใหม่ทุกๆ 10 เฟรมเพื่อลดโหลด)
        # ====================
        if frame_count % 10 == 0 or last_dashboard is None:
            dashboard = np.zeros((dash_h, dash_w, 3), dtype=np.uint8)
            dashboard[:] = BG_COLOR

            # Header
            cv2.rectangle(dashboard, (0, 0), (dash_w, 58), HEADER_COLOR, -1)
            draw_text_centered(dashboard, "VEHICLE COUNTER", dash_w // 2, 24, FONT, 0.65, ACCENT, 2)
            draw_text_centered(dashboard, datetime.now().strftime("%H:%M:%S"),
                               dash_w // 2, 46, FONT, 0.42, MUTED_COLOR, 1)

            cv2.rectangle(dashboard, (15, 65), (dash_w - 15, 105), DIVIDER_COLOR, -1)
            draw_text_centered(dashboard, f"TOTAL  {grand_total}", dash_w // 2, 92,
                               FONT, 0.72, (255, 255, 255), 2)

            cv2.line(dashboard, (15, 113), (dash_w - 15, 113), DIVIDER_COLOR, 1)

            y = 122
            for li in range(num_lines):
                lc  = LINE_COLORS[li]
                lbl = line_labels[li]
                cnt = counts_per_line[li]
                total_in  = sum(cnt["inbound"].values())
                total_out = sum(cnt["outbound"].values())
                line_total = total_in + total_out

                # Line section header
                cv2.rectangle(dashboard, (14, y), (dash_w - 14, y + 20), (28, 28, 45), -1)
                cv2.rectangle(dashboard, (14, y), (18, y + 20), lc, -1)
                cv2.putText(dashboard, f"  {lbl}", (20, y + 15),
                            FONT, 0.5, lc, 1, cv2.LINE_AA)
                total_str = f"= {line_total}"
                (tw, _), _ = cv2.getTextSize(total_str, FONT, 0.5, 2)
                cv2.putText(dashboard, total_str, (dash_w - 20 - tw, y + 15),
                            FONT, 0.5, (255, 255, 255), 2, cv2.LINE_AA)
                y += 24

                # Inbound row (compact tags)
                cv2.putText(dashboard, "  IN", (22, y + 12), FONT, 0.4, INBOUND_COLOR, 1, cv2.LINE_AA)
                xb = 60
                for cls_nm, count in cnt["inbound"].items():
                    tag = f"{cls_nm}:{count}"
                    (tw2, _), _ = cv2.getTextSize(tag, FONT, 0.38, 1)
                    cv2.rectangle(dashboard, (xb, y + 1), (xb + tw2 + 8, y + 15), (0, 80, 40), -1)
                    cv2.putText(dashboard, tag, (xb + 4, y + 12),
                                FONT, 0.38, INBOUND_COLOR, 1, cv2.LINE_AA)
                    xb += tw2 + 12
                y += 20

                # Outbound row (compact tags)
                cv2.putText(dashboard, "  OUT", (22, y + 12), FONT, 0.4, OUTBOUND_COLOR, 1, cv2.LINE_AA)
                xb = 66
                for cls_nm, count in cnt["outbound"].items():
                    tag = f"{cls_nm}:{count}"
                    (tw2, _), _ = cv2.getTextSize(tag, FONT, 0.38, 1)
                    cv2.rectangle(dashboard, (xb, y + 1), (xb + tw2 + 8, y + 15), (30, 20, 90), -1)
                    cv2.putText(dashboard, tag, (xb + 4, y + 12),
                                FONT, 0.38, OUTBOUND_COLOR, 1, cv2.LINE_AA)
                    xb += tw2 + 12
                y += 22

                # Divider between lines
                if li < num_lines - 1:
                    cv2.line(dashboard, (20, y), (dash_w - 20, y), DIVIDER_COLOR, 1)
                    y += 8

            # Buttons
            draw_rounded_rect(dashboard, (rld_x1, rld_y1), (rld_x2, rld_y2), BTN_RELOAD, radius=10)
            draw_text_centered(dashboard, "RELOAD STREAM",
                               (rld_x1 + rld_x2) // 2, rld_y1 + 20, FONT, 0.5, (255, 255, 255), 2)
            draw_text_centered(dashboard, "(or press  R)",
                               (rld_x1 + rld_x2) // 2, rld_y1 + 36, FONT, 0.35, (180, 220, 230))

            draw_rounded_rect(dashboard, (btn_x1, btn_y1), (btn_x2, btn_y2), BTN_STOP, radius=10)
            draw_text_centered(dashboard, "STOP & EXPORT",
                               (btn_x1 + btn_x2) // 2, btn_y1 + 20, FONT, 0.5, (255, 255, 255), 2)
            draw_text_centered(dashboard, "(or press  Q)",
                               (btn_x1 + btn_x2) // 2, btn_y1 + 36, FONT, 0.35, (180, 200, 255))

            mode_label = "FAST MODE" if fast_mode else "QUALITY MODE"
            mode_color = (0, 200, 100) if fast_mode else (200, 140, 0)
            cv2.circle(dashboard, (22, dash_h - 12), 5, mode_color, -1)
            cv2.putText(dashboard, mode_label, (32, dash_h - 7),
                        FONT, 0.35, mode_color, 1, cv2.LINE_AA)

            last_dashboard = dashboard
        cv2.imshow("Dashboard", last_dashboard)

        # Overlay มุมซ้ายบนบน frame (อันนี้อัปเดตทุกเฟรมได้ ไม่กินโหลดเท่าไหร่)
        ov = frame.copy()
        cv2.rectangle(ov, (0, 0), (210, 28), (10, 10, 20), -1)
        cv2.addWeighted(ov, 0.65, frame, 0.35, 0, frame)
        cv2.putText(frame, f"IN:{grand_in}  OUT:{grand_out}  TOT:{grand_total}",
                    (8, 20), FONT, 0.45, (0, 220, 255), 1, cv2.LINE_AA)

        cv2.imshow("Vehicle Counting", frame)

        key = cv2.waitKey(1) & 0xFF
        if key == ord('q') or key == 27:
            stop_counting = True
            break
        elif key == ord('r'):
            manual_reload = True

    if cap is not None:
        cap.release()
    cv2.destroyAllWindows()

    # ==========================================
    # ส่วนที่ 3: Export ผลลัพธ์
    # ==========================================
    # ส่วนที่ 3: Export ผลลัพธ์ (Excel เสมอ — Summary + Detail sheet)
    # ==========================================
    any_data = any(len(d) > 0 for d in export_data_per_line)

    if not any_data:
        messagebox.showinfo("Done", "นับเสร็จสิ้น — ไม่มีข้อมูลยานพาหนะที่บันทึก")
        return

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        messagebox.showerror("Error",
            "ต้องติดตั้ง openpyxl ก่อน:\n  pip install openpyxl")
        return

    try:
        wb = openpyxl.Workbook()

        hdr_font  = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill  = PatternFill("solid", fgColor="1C1C2E")
        line_fills = ["003040", "003A1A", "3A2000"]

        # --- Sheet: Summary ---
        ws_sum = wb.active
        ws_sum.title = "Summary"
        ws_sum.append(["Line", "Label", "Direction", "Class", "Count"])
        for cell in ws_sum[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

        for li in range(num_lines):
            lbl = line_labels[li]
            cnt = counts_per_line[li]
            for cls_nm, count in cnt["inbound"].items():
                ws_sum.append([f"Line {li+1}", lbl, "Inbound", cls_nm, count])
            for cls_nm, count in cnt["outbound"].items():
                ws_sum.append([f"Line {li+1}", lbl, "Outbound", cls_nm, count])

        ws_sum.column_dimensions["A"].width = 10
        ws_sum.column_dimensions["B"].width = 18
        ws_sum.column_dimensions["C"].width = 12
        ws_sum.column_dimensions["D"].width = 14
        ws_sum.column_dimensions["E"].width = 10

        # --- Sheet ต่อเส้น ---
        for li in range(num_lines):
            lbl = line_labels[li]
            sheet_name = f"Line{li+1}_{lbl}"[:31]
            ws = wb.create_sheet(title=sheet_name)

            ws.append(["Timestamp", "Vehicle_ID", "Class", "Direction"])
            fill_color = line_fills[li % len(line_fills)]
            for cell in ws[1]:
                cell.font = hdr_font
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.alignment = Alignment(horizontal="center")

            for row in export_data_per_line[li]:
                ws.append(row)

            ws.column_dimensions["A"].width = 22
            ws.column_dimensions["B"].width = 12
            ws.column_dimensions["C"].width = 14
            ws.column_dimensions["D"].width = 12

        wb.save(export_path)
        print(f"✅ บันทึก Excel: {export_path}")
        n_detail = num_lines
        messagebox.showinfo("Success",
            f"นับเสร็จสิ้น!\nข้อมูลบันทึกที่:\n{export_path}\n"
            f"(Summary + {n_detail} detail sheet{'s' if n_detail > 1 else ''})"
        )
    except Exception as e:
        print(f"❌ บันทึก Excel ไม่ได้: {e}")
        messagebox.showerror("Error", f"บันทึกไฟล์ Excel ไม่ได้:\n{e}")


# ==========================================
# GUI หน้าหลัก (Tkinter) — Dark Theme
# ==========================================
def open_gui():
    BG        = "#12121e"
    PANEL     = "#1c1c2e"
    CARD      = "#1e1e32"
    BORDER    = "#2e2e4a"
    ACCENT    = "#00c8ff"
    ACCENT2   = "#7b5ea7"
    BTN_GREEN = "#1db954"
    BTN_HOVER = "#17a046"
    TEXT      = "#e0e0ee"
    MUTED     = "#7070a0"
    ENTRY_BG  = "#161628"
    ENTRY_FG  = "#e0e0ee"

    FONT_TITLE  = ("Segoe UI", 16, "bold")
    FONT_SUBTITLE = ("Segoe UI", 8)
    FONT_HEAD   = ("Segoe UI", 9, "bold")
    FONT_BODY   = ("Segoe UI", 9)
    FONT_SMALL  = ("Segoe UI", 8)
    FONT_BTN    = ("Segoe UI", 11, "bold")
    FONT_LABEL  = ("Segoe UI", 8, "bold")

    DEFAULT_LINE_NAMES = ["Line 1", "Line 2", "Line 3"]

    root = tk.Tk()
    root.title("Vehicle Counter  —  YOLOv8")
    root.geometry("740x548")
    root.resizable(False, False)
    root.configure(bg=BG)

    source_type    = tk.IntVar(value=1)
    file_path_var  = tk.StringVar()
    stream_url_var = tk.StringVar(value="")
    export_var     = tk.StringVar(value="vehicle_counts.xlsx")
    fast_mode_var  = tk.BooleanVar(value=True)
    num_lines_var  = tk.IntVar(value=1)
    model_size_var  = tk.StringVar(value="yolov8n.pt")
    sensitivity_var = tk.StringVar(value="Normal (0.25)")
    line_name_vars  = [tk.StringVar(value=DEFAULT_LINE_NAMES[i]) for i in range(3)]

    def make_entry(parent, textvar, width=28, **kw):
        e = tk.Entry(parent, textvariable=textvar, bg=ENTRY_BG, fg=ENTRY_FG,
                     insertbackground=ACCENT, relief="flat", font=FONT_BODY, width=width,
                     highlightthickness=1, highlightbackground=BORDER,
                     highlightcolor=ACCENT, **kw)
        return e

    def make_label(parent, text, font=None, fg=None, **kw):
        return tk.Label(parent, text=text, font=font or FONT_BODY,
                        bg=CARD, fg=fg or TEXT, **kw)

    def card(parent, **kw):
        """Flat card with CARD background and border."""
        f = tk.Frame(parent, bg=CARD, highlightthickness=1, highlightbackground=BORDER, **kw)
        return f

    def section_title(parent, text):
        tk.Label(parent, text=text, font=FONT_LABEL, bg=CARD,
                 fg=ACCENT).pack(anchor="w", pady=(0, 6))

    def browse_file():
        path = filedialog.askopenfilename(
            title="เลือกไฟล์วิดีโอ",
            filetypes=[("Video Files", "*.mp4 *.avi *.mov *.mkv"), ("All Files", "*.*")]
        )
        if path:
            file_path_var.set(path)

    def on_num_lines_change(*_):
        n = num_lines_var.get()
        for i in range(3):
            if i < n:
                line_name_frames[i].pack(anchor="w", padx=4, pady=1)
            else:
                line_name_frames[i].pack_forget()

    def start_processing():
        if source_type.get() == 1:
            source = file_path_var.get().strip()
            if not source:
                messagebox.showwarning("Warning", "กรุณาเลือกไฟล์วิดีโอ")
                return
        else:
            raw = stream_url_var.get().strip()
            source = raw.split('\n')[0].split('\r')[0]
            if not source:
                messagebox.showwarning("Warning", "กรุณากรอก Streaming URL หรือ Camera ID (เช่น 0)")
                return

        n = num_lines_var.get()
        labels = [line_name_vars[i].get().strip() or DEFAULT_LINE_NAMES[i] for i in range(n)]
        export_path = export_var.get().strip() or "vehicle_counts.xlsx"
        fast_mode = fast_mode_var.get()
        model_size = model_size_var.get()

        sens_str = sensitivity_var.get()
        if "High" in sens_str:
            conf_th = 0.15
        elif "Low" in sens_str:
            conf_th = 0.40
        else:
            conf_th = 0.25
        iou_th = 0.45

        root.destroy()
        run_counter(source, export_path, fast_mode, n, labels, model_size, conf_th, iou_th)

    def on_enter_btn(e): start_btn.configure(bg=BTN_HOVER)
    def on_leave_btn(e): start_btn.configure(bg=BTN_GREEN)

    # ──────────────────────────────────────
    # HEADER
    # ──────────────────────────────────────
    header = tk.Frame(root, bg=PANEL, height=60)
    header.pack(fill="x")
    header.pack_propagate(False)

    hdr_left = tk.Frame(header, bg=PANEL)
    hdr_left.pack(side="left", padx=20, pady=8)
    tk.Label(hdr_left, text="🚗  VEHICLE COUNTER", font=FONT_TITLE,
             bg=PANEL, fg=ACCENT).pack(anchor="w")
    tk.Label(hdr_left, text="Real-time detection powered by YOLOv8", font=FONT_SUBTITLE,
             bg=PANEL, fg=MUTED).pack(anchor="w")

    tk.Label(header, text="v2.0", font=("Segoe UI", 8, "bold"),
             bg=PANEL, fg=ACCENT2).pack(side="right", padx=20)

    tk.Frame(root, bg=ACCENT, height=2).pack(fill="x")

    # ──────────────────────────────────────
    # BODY — two-column layout
    # ──────────────────────────────────────
    body = tk.Frame(root, bg=BG)
    body.pack(fill="both", expand=True, padx=14, pady=10)

    col_left  = tk.Frame(body, bg=BG)
    col_right = tk.Frame(body, bg=BG)
    col_left.pack(side="left", fill="both", expand=True, padx=(0, 7))
    col_right.pack(side="left", fill="both", expand=True, padx=(7, 0))

    # ── LEFT: Video Source ──────────────────
    src_card = card(col_left)
    src_card.pack(fill="x", pady=(0, 8))
    src_inner = tk.Frame(src_card, bg=CARD, padx=12, pady=10)
    src_inner.pack(fill="x")

    section_title(src_inner, "📂  VIDEO SOURCE")

    # File upload row
    rb_file = tk.Radiobutton(src_inner, text="Video File",
                   variable=source_type, value=1,
                   bg=CARD, fg=TEXT, selectcolor=PANEL,
                   activebackground=CARD, activeforeground=ACCENT,
                   font=FONT_BODY)
    rb_file.pack(anchor="w")

    file_frame = tk.Frame(src_inner, bg=CARD)
    file_frame.pack(fill="x", pady=(2, 6), padx=12)
    make_entry(file_frame, file_path_var, width=26).pack(side="left", ipady=4, padx=(0, 6), fill="x", expand=True)
    tk.Button(file_frame, text="Browse…", command=browse_file,
              bg=BORDER, fg=ACCENT, activebackground="#3a3a5c", activeforeground=ACCENT,
              relief="flat", font=FONT_SMALL, cursor="hand2", padx=8, pady=4).pack(side="left")

    # Stream row
    rb_stream = tk.Radiobutton(src_inner, text="Streaming URL / Camera ID",
                   variable=source_type, value=2,
                   bg=CARD, fg=TEXT, selectcolor=PANEL,
                   activebackground=CARD, activeforeground=ACCENT,
                   font=FONT_BODY)
    rb_stream.pack(anchor="w")

    stream_frame = tk.Frame(src_inner, bg=CARD)
    stream_frame.pack(fill="x", pady=(2, 4), padx=12)
    make_entry(stream_frame, stream_url_var, width=34).pack(side="left", ipady=4, fill="x", expand=True)

    # ── LEFT: Counting Lines ────────────────
    lines_card = card(col_left)
    lines_card.pack(fill="x", pady=(0, 8))
    lines_inner = tk.Frame(lines_card, bg=CARD, padx=12, pady=10)
    lines_inner.pack(fill="x")

    section_title(lines_inner, "📏  COUNTING LINES")

    lines_top = tk.Frame(lines_inner, bg=CARD)
    lines_top.pack(fill="x", pady=(0, 6))
    tk.Label(lines_top, text="จำนวนเส้นนับ:", font=FONT_BODY, bg=CARD, fg=MUTED).pack(side="left")
    num_lines_menu = tk.OptionMenu(lines_top, num_lines_var, 1, 2, 3,
                                   command=lambda _: on_num_lines_change())
    num_lines_menu.config(bg=BORDER, fg=TEXT, activebackground="#3a3a5c",
                          activeforeground=ACCENT, relief="flat",
                          highlightthickness=0, font=FONT_BODY, padx=6)
    num_lines_menu["menu"].config(bg=PANEL, fg=TEXT, activebackground=BORDER,
                                  activeforeground=ACCENT)
    num_lines_menu.pack(side="left", padx=8)

    LCOLORS = LINE_LABEL_COLORS_HEX
    line_name_frames = []
    for i in range(3):
        frm = tk.Frame(lines_inner, bg=CARD)
        line_name_frames.append(frm)
        dot_c = tk.Canvas(frm, width=10, height=10, bg=CARD, highlightthickness=0)
        dot_c.create_oval(1, 1, 9, 9, fill=LCOLORS[i], outline="")
        dot_c.pack(side="left", padx=(0, 5))
        tk.Label(frm, text=f"Line {i+1}:", font=FONT_BODY, bg=CARD, fg=TEXT).pack(side="left", padx=(0, 5))
        make_entry(frm, line_name_vars[i], width=16).pack(side="left", ipady=3)

    # Show only line 1 initially
    line_name_frames[0].pack(anchor="w", padx=4, pady=1)

    # ── LEFT: Export ────────────────────────
    exp_card = card(col_left)
    exp_card.pack(fill="x")
    exp_inner = tk.Frame(exp_card, bg=CARD, padx=12, pady=10)
    exp_inner.pack(fill="x")

    section_title(exp_inner, "📊  EXPORT")
    exp_row = tk.Frame(exp_inner, bg=CARD)
    exp_row.pack(fill="x")
    tk.Label(exp_row, text="ชื่อไฟล์ Excel:", font=FONT_BODY, bg=CARD, fg=MUTED).pack(side="left", padx=(0, 8))
    make_entry(exp_row, export_var, width=22).pack(side="left", ipady=4)

    # ── RIGHT: AI Settings ──────────────────
    ai_card = card(col_right)
    ai_card.pack(fill="x", pady=(0, 8))
    ai_inner = tk.Frame(ai_card, bg=CARD, padx=12, pady=10)
    ai_inner.pack(fill="x")

    section_title(ai_inner, "🤖  AI MODEL SETTINGS")

    # Model
    m_row = tk.Frame(ai_inner, bg=CARD)
    m_row.pack(fill="x", pady=(0, 6))
    tk.Label(m_row, text="Model:", font=FONT_BODY, bg=CARD, fg=MUTED, width=12, anchor="w").pack(side="left")
    model_menu = tk.OptionMenu(m_row, model_size_var, "yolov8n.pt", "yolov8s.pt", "yolov8m.pt")
    model_menu.config(bg=BORDER, fg=TEXT, activebackground="#3a3a5c",
                      activeforeground=ACCENT, relief="flat", highlightthickness=0,
                      font=FONT_BODY, padx=6)
    model_menu["menu"].config(bg=PANEL, fg=TEXT)
    model_menu.pack(side="left", padx=(0, 8))

    # Model hint
    tk.Label(ai_inner, text="  n = เร็ว  /  s = สมดุล  /  m = แม่นยำ (ช้า)",
             font=FONT_SMALL, bg=CARD, fg=MUTED).pack(anchor="w", pady=(0, 8))

    # Sensitivity
    s_row = tk.Frame(ai_inner, bg=CARD)
    s_row.pack(fill="x", pady=(0, 4))
    tk.Label(s_row, text="Sensitivity:", font=FONT_BODY, bg=CARD, fg=MUTED, width=12, anchor="w").pack(side="left")
    sens_menu = tk.OptionMenu(s_row, sensitivity_var,
                              "High (0.15) จับง่าย", "Normal (0.25)", "Low (0.40) ชัวร์เท่านั้น")
    sens_menu.config(bg=BORDER, fg=TEXT, activebackground="#3a3a5c",
                     activeforeground=ACCENT, relief="flat", highlightthickness=0,
                     font=FONT_BODY, padx=6)
    sens_menu["menu"].config(bg=PANEL, fg=TEXT)
    sens_menu.pack(side="left")

    tk.Label(ai_inner, text="  High = จับมอเตอร์ไซค์ได้มากขึ้น  /  Low = แม่นสูง",
             font=FONT_SMALL, bg=CARD, fg=MUTED).pack(anchor="w")

    # ── RIGHT: Performance ────────────────
    perf_card = card(col_right)
    perf_card.pack(fill="x", pady=(0, 8))
    perf_inner = tk.Frame(perf_card, bg=CARD, padx=12, pady=10)
    perf_inner.pack(fill="x")

    section_title(perf_inner, "⚡  PERFORMANCE")

    fast_cb = tk.Checkbutton(perf_inner,
                   text="Fast Mode  (320px — เร็วกว่า)",
                   variable=fast_mode_var,
                   bg=CARD, fg=TEXT, selectcolor=PANEL,
                   activebackground=CARD, activeforeground=ACCENT,
                   font=FONT_BODY)
    fast_cb.pack(anchor="w")
    tk.Label(perf_inner,
             text="  ปิด Fast Mode เพื่อจับรถเล็ก/มอเตอร์ไซค์ได้แม่นขึ้น",
             font=FONT_SMALL, bg=CARD, fg=MUTED).pack(anchor="w", pady=(2, 0))

    # ── RIGHT: How to use ────────────────
    tip_card = card(col_right)
    tip_card.pack(fill="x")
    tip_inner = tk.Frame(tip_card, bg=CARD, padx=12, pady=10)
    tip_inner.pack(fill="x")

    section_title(tip_inner, "💡  วิธีใช้งาน")
    tips = [
        "① เลือกแหล่งวิดีโอ (ไฟล์ หรือ URL)",
        "② ตั้งจำนวน + ชื่อเส้นนับ",
        "③ กด Start แล้ววาดเส้นบนภาพ",
        "④ กด  C  ยืนยันเส้น,  Q  หยุด",
    ]
    for t in tips:
        tk.Label(tip_inner, text=t, font=FONT_SMALL, bg=CARD,
                 fg=MUTED, anchor="w").pack(fill="x")

    # ──────────────────────────────────────
    # FOOTER — START BUTTON
    # ──────────────────────────────────────
    footer = tk.Frame(root, bg=BG)
    footer.pack(fill="x", padx=14, pady=(4, 12))

    start_btn = tk.Button(footer, text="🚀  เริ่มนับยานพาหนะ  (Start Counting)",
                          font=FONT_BTN, bg=BTN_GREEN, fg="white",
                          activebackground=BTN_HOVER, activeforeground="white",
                          relief="flat", cursor="hand2", padx=20, pady=11,
                          command=start_processing)
    start_btn.pack(fill="x")
    start_btn.bind("<Enter>", on_enter_btn)
    start_btn.bind("<Leave>", on_leave_btn)

    root.mainloop()


if __name__ == "__main__":
    open_gui()
